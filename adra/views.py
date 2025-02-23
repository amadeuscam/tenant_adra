import glob
import io
import os
import subprocess  # imported for using queue.Empty exception
from datetime import datetime
from multiprocessing import Process
from pathlib import Path

import telegram
from allauth.account.adapter import DefaultAccountAdapter
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.messages.views import SuccessMessageMixin
from django.core.files.storage import FileSystemStorage
from django.db import connection
from django.db.models import Q, Sum
from django.http import HttpResponse, HttpResponseNotFound, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.decorators.cache import cache_page, never_cache
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from django_tenants.utils import get_tenant_model
from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter
from rest_framework import permissions, viewsets
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

from adra.utils.adra_util import (
    AdraUtils,
    AgeCalculacion,
    DeliverySheet,
    RecuntoBeneficiarios,
    UploadExcelUsers,
    ValoracionSocial,
)
from delegaciones.models import BeneficiariosGlobales, Delegaciones

from .filters import AlimentosFilters
from .forms import (
    AlimentosFrom,
    AlimentosRepatrirForm,
    AlmacenAlimentosFrom,
    DelegacionForm,
    HijoForm,
    PersonaForm,
    ProfileEditForm,
    UserEditForm,
)
from .models import (
    Alimentos,
    AlimentosRepatir,
    AlmacenAlimentos,
    Hijo,
    Persona,
)
from .serializers import (
    AlacenAlimentosSerializer,
    PersonaSerializer,
    UserSerializer,
)


@login_required
def edit(request):
    if request.method == "POST":
        user_form = UserEditForm(instance=request.user, data=request.POST)
        profile_form = ProfileEditForm(
            instance=request.user.profile,
            data=request.POST,
            files=request.FILES,
        )
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(
                request, "El perfil se ha actualizado correctamente"
            )
            return redirect("adra:edit-profile")
        else:
            messages.error(request, "Error updating your profile")
    else:
        user_form = UserEditForm(instance=request.user)
        profile_form = ProfileEditForm(instance=request.user.profile)
    return render(
        request,
        "account/edit-perfile.html",
        {"user_form": user_form, "profile_form": profile_form},
    )


class PersonaListView(LoginRequiredMixin, ListView):
    template_name = "adra/index.html"
    context_object_name = "beneficiarios"
    model = Persona
    login_url = "account_login"
    paginate_by = 12
    queryset = Persona.objects.filter(active=True).exclude(covid=True)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # enviar al misma pagina varios contextos
        context["hijo"] = Hijo.objects
        context["hijomenor"] = Hijo
        context["nbar"] = "home"
        context["debug"] = settings.DEBUG
        context["platform_name"] = settings.PLATFORM_NAME

        return context


# ir a la pagina detallada de cada persona
class PersonaDetailView(LoginRequiredMixin, DetailView):
    model = Persona
    template_name = "adra/detail.html"
    context_object_name = "persona"
    login_url = "account_login"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = AlimentosFilters(
            self.request.GET, queryset=self.get_queryset()
        )
        return context


class PersonaCreateView(LoginRequiredMixin, CreateView):
    model = Persona
    success_url = "/"
    form_class = PersonaForm

    def form_valid(self, form):
        # print(form)
        # print(form.cleaned_data)
        logic = Q(
            documentacion_beneficiario=form.cleaned_data["dni"]
            if form.cleaned_data["dni"]
            else form.cleaned_data["otros_documentos"],
            telefono=form.cleaned_data["telefono"],
            # nombre_beneficiario=form.cleaned_data["nombre_apellido"],
            _connector=Q.OR,
        )
        beneficario_global = BeneficiariosGlobales.objects.filter(
            logic, nombre_beneficiario=form.cleaned_data["nombre_apellido"]
        )
        # print(beneficario_global)
        # print(beneficario_global.count())

        if beneficario_global.count() > 0:
            for beneficario in beneficario_global:
                # print(beneficario.delegacion_name)

                messages.warning(
                    self.request,
                    "{} es beneficiario de la oar {} desde {}".format(
                        beneficario.nombre_beneficiario,
                        beneficario.delegacion_name,
                        beneficario.created_at.strftime("%d %B, %Y"),
                    ),
                )
            return super().form_invalid(form)
        else:
            form.instance.modificado_por = self.request.user
            messages.add_message(
                self.request,
                messages.SUCCESS,
                "Beneficiario añadido con exito!",
            )
            return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["nbar"] = "create"

        try:
            p = Persona.objects.latest("created_at")
            context["bas"] = p
        except Persona.DoesNotExist:
            context["bas"] = None

        return context


class PersonaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Persona
    form_class = PersonaForm

    success_message = "Beneficiario actualizado satisfactoriamente!!"

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["up"] = "update"
        return context


class PersonaDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Persona
    success_message = "Beneficirio borrado corectamente"
    success_url = "/"


def adauga_alimentos_persona(request, pk):
    persona = get_object_or_404(Persona, pk=pk)
    almacen = AlmacenAlimentos.objects.all().first()

    if request.method == "POST":
        print(request.POST)
        a_form = AlimentosFrom(request.POST)
        if a_form.is_valid():
            alimentos = a_form.save(commit=False)

            almacen.alimento_1 -= alimentos.alimento_1
            almacen.alimento_2 -= alimentos.alimento_2
            almacen.alimento_3 -= alimentos.alimento_3
            almacen.alimento_4 -= alimentos.alimento_4
            almacen.alimento_5 -= alimentos.alimento_5
            almacen.alimento_6 -= alimentos.alimento_6
            almacen.alimento_7 -= alimentos.alimento_7
            almacen.alimento_8 -= alimentos.alimento_8
            almacen.alimento_9 -= alimentos.alimento_9
            almacen.alimento_10 -= alimentos.alimento_10
            almacen.alimento_11 -= alimentos.alimento_11
            almacen.alimento_12 -= alimentos.alimento_12
            almacen.alimento_13 -= alimentos.alimento_13
            almacen.alimento_14 -= alimentos.alimento_14
            almacen.alimento_15 -= alimentos.alimento_15
            almacen.alimento_16 -= alimentos.alimento_16

            alimentos.persona = persona
            alimentos.modificado_por = request.user
            almacen.save()
            alimentos.save()
            messages.success(
                request, "Los alimentos se han entregado correctamente!"
            )
            return redirect(persona)

    else:
        alm_repatir = AlimentosRepatir.objects.all().first()

        init_reparto_alimento = {}
        for index in range(1, 17):
            if getattr(alm_repatir, f"alimento_{index}_type") == "unidad":
                babys = len(
                    [fam for fam in persona.hijo.all() if fam.age <= 3]
                )
                if babys > 0 and index in [12, 13]:
                    init_reparto_alimento[f"alimento_{index}"] = (
                        getattr(alm_repatir, f"alimento_{index}") * babys
                    )
                    continue
                else:
                    if index in [12, 13]:
                        init_reparto_alimento[f"alimento_{index}"] = 0
                        continue
                init_reparto_alimento[f"alimento_{index}"] = (
                    getattr(alm_repatir, f"alimento_{index}")
                    * persona.numero_beneficiarios
                )
            else:
                if 0 <= persona.numero_beneficiarios <= 3:
                    init_reparto_alimento[f"alimento_{index}"] = getattr(
                        alm_repatir, f"alimento_{index}_0_3"
                    )
                elif 4 <= persona.numero_beneficiarios <= 6:
                    init_reparto_alimento[f"alimento_{index}"] = getattr(
                        alm_repatir, f"alimento_{index}_4_6"
                    )
                elif 7 <= persona.numero_beneficiarios <= 10:
                    init_reparto_alimento[f"alimento_{index}"] = getattr(
                        alm_repatir, f"alimento_{index}_7_9"
                    )
                else:
                    print("ninguna es correcta")
        init_reparto_alimento["alimento_4"] = 0
        print(init_reparto_alimento)  # no tocar, es para los test esto
        a_form = AlimentosFrom(initial=init_reparto_alimento)
    return render(request, "adra/alimentos_form.html", {"form": a_form})


class PersonaAlimentosUpdateView(LoginRequiredMixin, UpdateView):
    model = Alimentos
    fields = [
        "alimento_1",
        "alimento_2",
        "alimento_3",
        "alimento_4",
        "alimento_5",
        "alimento_6",
        "alimento_7",
        "alimento_8",
        "alimento_9",
        "alimento_10",
        "alimento_11",
        "alimento_12",
        "alimento_13",
        "alimento_14",
        "alimento_15",
        "alimento_16",
        "fecha_recogida",
        "signature",
    ]

    def form_valid(self, form):
        clean = form.changed_data
        # print(clean)
        # print(form)
        almacen = AlmacenAlimentos.objects.all().first()

        if "alimento_1" in clean:
            valor_anterior_alimento_1 = form.initial["alimento_1"]
            if form.instance.alimento_1 > valor_anterior_alimento_1:
                restante = abs(
                    form.instance.alimento_1 - valor_anterior_alimento_1
                )
                almacen.alimento_1 -= restante
            else:
                restante = abs(
                    form.instance.alimento_1 - valor_anterior_alimento_1
                )
                almacen.alimento_1 += restante

        if "alimento_2" in clean:
            valor_anterior_alimento_2 = form.initial["alimento_2"]
            if form.instance.alimento_2 > valor_anterior_alimento_2:
                restante = abs(
                    form.instance.alimento_2 - valor_anterior_alimento_2
                )
                almacen.alimento_2 -= restante
            else:
                restante = abs(
                    form.instance.alimento_2 - valor_anterior_alimento_2
                )
                almacen.alimento_2 += restante

        if "alimento_3" in clean:
            valor_anterior_alimento_3 = form.initial["alimento_3"]
            if form.instance.alimento_3 > valor_anterior_alimento_3:
                restante = abs(
                    form.instance.alimento_3 - valor_anterior_alimento_3
                )
                almacen.alimento_3 -= restante
            else:
                restante = abs(
                    form.instance.alimento_3 - valor_anterior_alimento_3
                )
                almacen.alimento_3 += restante

        if "alimento_4" in clean:
            valor_anterior_alimento_4 = form.initial["alimento_4"]
            if form.instance.alimento_4 > valor_anterior_alimento_4:
                restante = abs(
                    form.instance.alimento_4 - valor_anterior_alimento_4
                )
                almacen.alimento_4 -= restante
            else:
                restante = abs(
                    form.instance.alimento_4 - valor_anterior_alimento_4
                )
                almacen.alimento_4 += restante

        if "alimento_5" in clean:
            valor_anterior_alimento_5 = form.initial["alimento_5"]
            if form.instance.alimento_5 > valor_anterior_alimento_5:
                restante = abs(
                    form.instance.alimento_5 - valor_anterior_alimento_5
                )
                almacen.alimento_5 -= restante
            else:
                restante = abs(
                    form.instance.alimento_5 - valor_anterior_alimento_5
                )
                almacen.alimento_5 += restante

        if "alimento_6" in clean:
            valor_anterior_alimento_6 = form.initial["alimento_6"]
            if form.instance.alimento_6 > valor_anterior_alimento_6:
                restante = abs(
                    form.instance.alimento_6 - valor_anterior_alimento_6
                )
                almacen.alimento_6 -= restante
            else:
                restante = abs(
                    form.instance.alimento_6 - valor_anterior_alimento_6
                )
                almacen.alimento_6 += restante

        if "alimento_7" in clean:
            valor_anterior_alimento_7 = form.initial["alimento_7"]
            if form.instance.alimento_7 > valor_anterior_alimento_7:
                restante = abs(
                    form.instance.alimento_7 - valor_anterior_alimento_7
                )
                almacen.alimento_7 -= restante
            else:
                restante = abs(
                    form.instance.alimento_7 - valor_anterior_alimento_7
                )
                almacen.alimento_7 += restante

        if "alimento_8" in clean:
            valor_anterior_alimento_8 = form.initial["alimento_8"]
            if form.instance.alimento_8 > valor_anterior_alimento_8:
                restante = abs(
                    form.instance.alimento_8 - valor_anterior_alimento_8
                )
                almacen.alimento_8 -= restante
            else:
                restante = abs(
                    form.instance.alimento_8 - valor_anterior_alimento_8
                )
                almacen.alimento_8 += restante

        if "alimento_9" in clean:
            valor_anterior_alimento_9 = form.initial["alimento_9"]
            if form.instance.alimento_9 > valor_anterior_alimento_9:
                restante = abs(
                    form.instance.alimento_9 - valor_anterior_alimento_9
                )
                almacen.alimento_9 -= restante
            else:
                restante = abs(
                    form.instance.alimento_9 - valor_anterior_alimento_9
                )
                almacen.alimento_9 += restante

        if "alimento_10" in clean:
            valor_anterior_alimento_10 = form.initial["alimento_10"]
            if form.instance.alimento_10 > valor_anterior_alimento_10:
                restante = abs(
                    form.instance.alimento_10 - valor_anterior_alimento_10
                )
                almacen.alimento_10 -= restante
            else:
                restante = abs(
                    form.instance.alimento_10 - valor_anterior_alimento_10
                )
                almacen.alimento_10 += restante

        if "alimento_11" in clean:
            valor_anterior_alimento_11 = form.initial["alimento_11"]
            if form.instance.alimento_11 > valor_anterior_alimento_11:
                restante = abs(
                    form.instance.alimento_11 - valor_anterior_alimento_11
                )
                almacen.alimento_11 -= restante
            else:
                restante = abs(
                    form.instance.alimento_11 - valor_anterior_alimento_11
                )
                almacen.alimento_11 += restante

        if "alimento_12" in clean:
            valor_anterior_alimento_12 = form.initial["alimento_12"]
            if form.instance.alimento_12 > valor_anterior_alimento_12:
                restante = abs(
                    form.instance.alimento_12 - valor_anterior_alimento_12
                )
                almacen.alimento_12 -= restante
            else:
                restante = abs(
                    form.instance.alimento_12 - valor_anterior_alimento_12
                )
                almacen.alimento_12 += restante
                
        if "alimento_13" in clean:
            valor_anterior_alimento_13 = form.initial["alimento_13"]
            if form.instance.alimento_13 > valor_anterior_alimento_13:
                restante = abs(
                    form.instance.alimento_13 - valor_anterior_alimento_13
                )
                almacen.alimento_13 -= restante
            else:
                restante = abs(
                    form.instance.alimento_13 - valor_anterior_alimento_13
                )
                almacen.alimento_13 += restante

        if "alimento_14" in clean:
            valor_anterior_alimento_14 = form.initial["alimento_14"]
            if form.instance.alimento_14 > valor_anterior_alimento_14:
                restante = abs(
                    form.instance.alimento_14 - valor_anterior_alimento_14
                )
                almacen.alimento_14 -= restante
            else:
                restante = abs(
                    form.instance.alimento_14 - valor_anterior_alimento_14
                )
                almacen.alimento_14 += restante


        if "alimento_15" in clean:
            valor_anterior_alimento_15 = form.initial["alimento_15"]
            if form.instance.alimento_15 > valor_anterior_alimento_15:
                restante = abs(
                    form.instance.alimento_15 - valor_anterior_alimento_15
                )
                almacen.alimento_15 -= restante
            else:
                restante = abs(
                    form.instance.alimento_15 - valor_anterior_alimento_15
                )
                almacen.alimento_15 += restante
                
        if "alimento_16" in clean:
            valor_anterior_alimento_16 = form.initial["alimento_16"]
            if form.instance.alimento_16 > valor_anterior_alimento_16:
                restante = abs(
                    form.instance.alimento_16 - valor_anterior_alimento_16
                )
                almacen.alimento_16 -= restante
            else:
                restante = abs(
                    form.instance.alimento_16 - valor_anterior_alimento_16
                )
                almacen.alimento_16 += restante
                

        almacen.save()
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)


class PersonaAlimentosDeleteView(LoginRequiredMixin, DeleteView):
    model = Alimentos
    success_url = "/"


class AlmacenListView(LoginRequiredMixin, ListView):
    model = AlmacenAlimentos
    template_name = "adra_almacen/index.html"
    context_object_name = "almacen_adra"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["nbar"] = "almacen_a"
        return context


class AlmacenUpdateView(LoginRequiredMixin, UpdateView):
    model = AlmacenAlimentos
    form_class = AlmacenAlimentosFrom

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["nbar"] = "almacen_a"
        return context


def anadir_familiar(request, pk):
    persona = get_object_or_404(Persona, pk=pk)
    if request.method == "POST":
        # print(request.POST)
        h_form = HijoForm(request.POST)
        if h_form.is_valid():
            hijo = h_form.save(commit=False)
            hijo.persona = persona
            hijo.modificado_por = request.user
            hijo.save()
            return redirect(persona)

    else:
        h_form = HijoForm()
    return render(request, "adra/hijo_form.html", {"form": h_form})


class HijoUpdateView(LoginRequiredMixin, UpdateView):
    model = Hijo
    form_class = HijoForm

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["up"] = "update"
        return context


class HijoDeleteView(LoginRequiredMixin, DeleteView):
    model = Hijo
    persoan = Persona

    # success_url = reverse_lazy("adra:persona-home")

    def get_success_url(self):
        return reverse_lazy(
            "adra:personas_detail", kwargs={"pk": self.object.persona.pk}
        )


@login_required
def buscar(request):
    if "q" in request.GET and request.GET["q"]:
        q = request.GET["q"]

        if q.isdigit():
            ultima_persona = Persona.objects.filter(
                Q(numero_adra=q) | Q(telefono=q)
            ).filter(active=True)
        else:
            ultima_persona = Persona.objects.filter(
                nombre_apellido__icontains=q
            ).filter(active=True)
        if ultima_persona:
            return render(
                request,
                "adra/index.html",
                {"beneficiarios": ultima_persona, "query": q},
            )
        else:
            beneficiarios_no_activos = Persona.objects.filter(
                active=False
            ).order_by("-numero_adra")
            return render(
                request,
                "adra/no_beneficiarios.html",
                {"ben": beneficiarios_no_activos},
            )
    else:
        return HttpResponse("<h1>Por favor buscar con otro criterio</h1>")


@login_required
def statistics_persona(request):
    beneficiar = Persona.objects.filter(active=True).exclude(covid=True)

    familiares = Hijo.objects.filter(
        persona__in=Persona.objects.filter(active=True).exclude(covid=True)
    )

    ages = AgeCalculacion(beneficiar, familiares).calculate_age()

    ages.update({"nbar": "stat"})

    return render(request, "statistics/index.html", ages)


def telegram_messages(request):
    """
    Es encargado de mandar mensajes al grupo de telegram

    Args:
        request (request): se recibice el mensaje y el grupo de personas

    Returns:
        HttpResponse
    """

    if request.method == "POST":
        dom = str(request.POST.get("dom_select"))
        mensaje_propio = request.POST.get("mensaje_propio")
        if dom and mensaje_propio:
            bot = telegram.Bot(token=str(os.getenv("TELEGRAM_TOKEN")))
            persona = None

            if dom == "todos":
                persona = Persona.objects.filter(active=True).exclude(
                    covid=True
                )
            else:
                persona = (
                    Persona.objects.filter(active=True)
                    .filter(
                        Q(categoria=f"Domingo {int(dom)}")
                        | Q(categoria=int(dom)),
                        ciudad__icontains="Torrejon de ardoz",
                    )
                    .exclude(covid=True)
                )

            per_list = [p.nombre_apellido for p in persona]
            list_format = "\n".join(per_list)

            bot.send_message(
                str(os.getenv("TELEGRAM_CANAL")),
                f"*{list_format}*",
                parse_mode=telegram.ParseMode.MARKDOWN_V2,
            )
            bot.send_message(
                str(os.getenv("TELEGRAM_CANAL")),
                f"*{mensaje_propio}*",
                parse_mode=telegram.ParseMode.MARKDOWN_V2,
            )

            messages.success(request, "El mensaje se ha mandado correctamente")
            return render(request, "telegram/index.html", {"nbar": "tel"})
    else:
        return render(request, "telegram/index.html", {"nbar": "tel"})


def export_users_csv(request):
    beneficiarios_queryset = Persona.objects.order_by("numero_adra").filter(
        active=True
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.\
            spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=beneficiarios.xlsx"

    # workbook = Workbook()
    workbook = load_workbook(
        filename="source_files/Listado_beneficiarios.xlsx"
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    # Delete the default worksheet
    # workbook.remove(workbook.active)
    worksheet.title = "Beneficiarios"
    # Define the titles for columns
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 40
    worksheet.column_dimensions["C"].width = 30
    worksheet.column_dimensions["E"].width = 40
    # columns = [
    #     "Numar adra",
    #     "Nombre
    #     "Representante familiar",
    #     "Dni",
    #     "Pasaporte",
    #     "Fecha de nacimiento",
    # ]
    row_num = 6
    fill = PatternFill(start_color="43fb32", fill_type="solid")
    # Assign the titles for each cell of the header
    # for col_num, column_title in enumerate(columns, 1):
    #     cell = worksheet.cell(row=row_num, column=col_num)
    #     cell.value = column_title
    from openpyxl.styles import Font

    fontStyle = Font(size="12")
    count = 0
    for ben in beneficiarios_queryset:
        row_num += 1
        count += 1

        # Define the data for each cell in the row
        row = [
            f"{count}-{ben.numero_adra}",
            ben.nombre_apellido,
            1,
            ben.dni,
            ben.otros_documentos,
            ben.fecha_nacimiento.strftime("%d/%m/%Y"),
        ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(
                row=row_num, column=col_num, value=cell_value
            )
            # cell.value = cell_value
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.font = fontStyle

        for d in ben.hijo.filter(active=True):
            count += 1
            row_hijos = [
                f"{count}",
                d.nombre_apellido,
                "",
                d.dni,
                d.otros_documentos,
                d.fecha_nacimiento.strftime("%d/%m/%Y"),
            ]
            row_num += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row_hijos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal="center")
                cell.font = fontStyle

    workbook.save(response)

    return response


@login_required
@cache_page(60 * 15)
def buscar_fecha(request):
    # print(request.GET['download'] )
    alimentos_list = Alimentos.objects.all()
    # print(request.GET)
    user_filter = AlimentosFilters(request.GET, queryset=alimentos_list)
    alimento_1 = user_filter.qs.aggregate(Sum("alimento_1"))
    alimento_2 = user_filter.qs.aggregate(Sum("alimento_2"))
    alimento_3 = user_filter.qs.aggregate(Sum("alimento_3"))
    alimento_4 = user_filter.qs.aggregate(Sum("alimento_4"))
    alimento_5 = user_filter.qs.aggregate(Sum("alimento_5"))
    alimento_6 = user_filter.qs.aggregate(Sum("alimento_6"))
    alimento_7 = user_filter.qs.aggregate(Sum("alimento_7"))
    alimento_8 = user_filter.qs.aggregate(Sum("alimento_8"))
    alimento_9 = user_filter.qs.aggregate(Sum("alimento_9"))
    alimento_10 = user_filter.qs.aggregate(Sum("alimento_10"))
    alimento_11 = user_filter.qs.aggregate(Sum("alimento_11"))
    alimento_12 = user_filter.qs.aggregate(Sum("alimento_12"))
    alimento_13 = user_filter.qs.aggregate(Sum("alimento_13"))
    alimento_14 = user_filter.qs.aggregate(Sum("alimento_14"))
    alimento_15 = user_filter.qs.aggregate(Sum("alimento_15"))
    alimento_16 = user_filter.qs.aggregate(Sum("alimento_16"))

    if "download" in request.POST:
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.\
            spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = "attachment; filename=alimentos_repartidos.xlsx"  # noqa

        workbook = Workbook()
        # Get active worksheet/tab
        worksheet = workbook.active
        # Delete the default worksheet
        # workbook.remove(workbook.active)
        worksheet.title = "Alimentos Repartidos"
        # Define the titles for columns
        columns = [
            settings.ALIMENTOS_METADATA["alimento_1"]["name"],
            settings.ALIMENTOS_METADATA["alimento_2"]["name"],
            settings.ALIMENTOS_METADATA["alimento_3"]["name"],
            settings.ALIMENTOS_METADATA["alimento_4"]["name"],
            settings.ALIMENTOS_METADATA["alimento_5"]["name"],
            settings.ALIMENTOS_METADATA["alimento_6"]["name"],
            settings.ALIMENTOS_METADATA["alimento_7"]["name"],
            settings.ALIMENTOS_METADATA["alimento_8"]["name"],
            settings.ALIMENTOS_METADATA["alimento_9"]["name"],
            settings.ALIMENTOS_METADATA["alimento_10"]["name"],
            settings.ALIMENTOS_METADATA["alimento_11"]["name"],
            settings.ALIMENTOS_METADATA["alimento_12"]["name"],
            settings.ALIMENTOS_METADATA["alimento_13"]["name"],
            settings.ALIMENTOS_METADATA["alimento_14"]["name"],
            settings.ALIMENTOS_METADATA["alimento_15"]["name"],
            settings.ALIMENTOS_METADATA["alimento_16"]["name"],
            "Fecha Recogida",
            "Beneficiario",
        ]

        row_num = 1
        # fill = PatternFill(start_color="43fb32", fill_type="solid")
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        from openpyxl.styles import Font

        fontStyle = Font(size="12")
        count = 0
        for alimento in alimentos_list:
            row_num += 1
            # Define the data for each cell in the row
            row = [
                alimento.alimento_1,
                alimento.alimento_2,
                alimento.alimento_3,
                alimento.alimento_4,
                alimento.alimento_5,
                alimento.alimento_6,
                alimento.alimento_7,
                alimento.alimento_8,
                alimento.alimento_9,
                alimento.alimento_10,
                alimento.alimento_11,
                alimento.alimento_12,
                alimento.alimento_13,
                alimento.alimento_14,
                alimento.alimento_15,
                alimento.alimento_16,
                alimento.fecha_recogida.strftime("%d/%m/%Y"),
                alimento.persona.nombre_apellido,
            ]
            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(
                    row=row_num, column=col_num, value=cell_value
                )
                # cell.value = cell_value
                # cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
                cell.font = fontStyle

        # auto resize colum width
        for col in worksheet.columns:
            print(dir(col))
            print(col)
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(response)

        return response

    return render(
        request,
        "busqueda_a/view.html",
        {
            "filter": user_filter,
            "alimento_1": alimento_1,
            "alimento_2": alimento_2,
            "alimento_3": alimento_3,
            "alimento_4": alimento_4,
            "alimento_5": alimento_5,
            "alimento_6": alimento_6,
            "alimento_7": alimento_7,
            "alimento_8": alimento_8,
            "alimento_9": alimento_9,
            "alimento_10": alimento_10,
            "alimento_11": alimento_11,
            "alimento_12": alimento_12,
            "alimento_13": alimento_13,
            "alimento_14": alimento_14,
            "alimento_15": alimento_15,
            "alimento_16": alimento_16,
            "nbar": "buscar_av",
        },
    )


def generar_hoja_entrega(request, pk, mode):
    """
    Generador de hoja de entrega para cada beneficiario en parte
    :param request:
    :param pk: id persona
    :return: pdf generado

    Args:
        mode:
    """
    # print("mode generate->", mode)
    # print(request.tenant)
    # print(request.tenant.oar)
    tenant_info = request.tenant
    persona = Persona.objects.get(id=pk)

    alimentos = persona.alimentos.all().order_by("fecha_recogida")
    pdf = DeliverySheet(persona, tenant_info).add_signature(alimentos)
    response = HttpResponse(content_type="application/pdf")
    response[
        "Content-Disposition"
    ] = f"attachment; filename={persona.numero_adra}.pdf"
    pdf.write(response)
    return response


def set_need_appearances_writer(writer):
    from PyPDF2.generic import (
        BooleanObject,
        IndirectObject,
        NameObject,
        NumberObject,
        TextStringObject,
    )

    """
    Helper para escribir el pdf
    :param writer: el pdf
    :return:
    """
    try:
        catalog = writer._root_object
        # get the AcroForm tree and add "/NeedAppearances attribute
        if "/AcroForm" not in catalog:
            writer._root_object.update(
                {
                    NameObject("/AcroForm"): IndirectObject(
                        len(writer._objects), 0, writer
                    )
                }
            )

        need_appearances = NameObject("/NeedAppearances")
        writer._root_object["/AcroForm"][need_appearances] = BooleanObject(
            True
        )
        return writer

    except Exception as e:
        print("set_need_appearances_writer() catch : ", repr(e))
        return writer


def combine_word_documents(files):
    merged_document = Document()

    for index, file in enumerate(files):
        sub_doc = Document(file)

        # Don't add a page break if you've reached the last file.
        if index < len(files) - 1:
            sub_doc.add_page_break()

        for element in sub_doc.element.body:
            merged_document.element.body.append(element)

    return merged_document


@never_cache
def generar_hoja_entrega_global(request):
    tenant_info = request.tenant
    beneficiarios = Persona.objects.filter(active=True).exclude(covid=True)
    pdf_writer = PdfFileWriter()
    set_need_appearances_writer(pdf_writer)
    res_pdf = io.BytesIO()
    for beneficiar in beneficiarios:
        buf = io.BytesIO()
        dss = DeliverySheet(
            beneficiar, tenant_info
        ).add_signature_all_beneficiarios()
        dss.write(buf)
        buf.seek(0)

        reader = PdfFileReader(stream=buf)
        for pag in range(0, reader.getNumPages()):
            pdf_writer.addPage(reader.getPage(pag))

    pdf_writer.write(res_pdf)
    response = HttpResponse(res_pdf.getvalue(), content_type="application/pdf")
    current_date = datetime.today().strftime("%d-%m-%Y %H:%M:%S")
    response[
        "Content-Disposition"
    ] = f'attachment; filename="todas_las_entregas_{current_date}.pdf"'
    return response


@never_cache
def valoracion_social_global(request):
    beneficiarios = (
        Persona.objects.filter(active=True)
        .exclude(covid=True)
        .order_by("-numero_adra")
    )

    val_lst = []
    for beneficiar in beneficiarios:
        valoracion_social_bt = io.BytesIO()
        val = ValoracionSocial(beneficiar).get_valoracion()
        val.write(valoracion_social_bt)
        valoracion_social_bt.seek(0)
        val_lst.append(valoracion_social_bt)

    merged_document = combine_word_documents(val_lst)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"  # noqa
    )
    current_date = datetime.today().strftime("%d-%m-%Y %H:%M:%S")
    response[
        "Content-Disposition"
    ] = f"attachment; filename=hojas_valoracion_all_{current_date}.docx"
    merged_document.save(response)
    return response


@never_cache
def generar_hoja_valoracion_social(request, pk):
    persona = Persona.objects.get(pk=pk, active=True)
    val = ValoracionSocial(persona).get_valoracion()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"  # noqa
    )
    response[
        "Content-Disposition"
    ] = f"attachment; filename={persona.numero_adra}.docx"
    val.write(response)

    return response


class PersonaViewSet(viewsets.ModelViewSet):
    """
    This viewset automatically provides `list`, `create`, `retrieve`,
    `update` and `destroy` actions.
    """

    queryset = Persona.objects.all().order_by("-numero_adra")
    serializer_class = PersonaSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_fields = ("numero_adra", "nombre_apellido")

    def get_queryset(self):
        numero_adra = self.request.query_params.get("numero_adra", None)
        nombre_apellido = self.request.query_params.get(
            "nombre_apellido", None
        )
        queryset = None
        if numero_adra is not None:
            queryset = Persona.objects.filter(numero_adra=numero_adra)
        if nombre_apellido is not None:
            queryset = Persona.objects.filter(
                nombre_apellido__icontains=nombre_apellido
            )
        else:
            queryset = Persona.objects.all().order_by("-numero_adra")
        return queryset

    # def perform_destroy(self, request, *args, **kwargs):
    #         try:
    #             instance = self.get_object()
    #             print(instance)
    #             self.perform_destroy(instance)
    #         except Http404:
    #             pass
    #         return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_create(self, serializer):
        serializer.save(modificado_por=self.request.user)


class AlmacenViewSet(viewsets.ReadOnlyModelViewSet):
    """
    This viewset automatically provides `list`, `create`, `retrieve`,
    `update` and `destroy` actions.
    """

    queryset = AlmacenAlimentos.objects.all()
    serializer_class = AlacenAlimentosSerializer
    permission_classes = [permissions.IsAuthenticated]


class UserViewSet(viewsets.ModelViewSet):
    """
    This viewset automatically provides `list` and `detail` actions.
    """

    queryset = get_user_model().objects.filter(is_superuser=True)
    serializer_class = UserSerializer


@login_required
def get_data(request):
    """ " manda datos a los graficos de echats"""
    count_altas = []
    years = Persona.objects.dates("created_at", "year").values_list(
        "created_at__year", flat=True
    )
    # print(years)
    for year in years:
        count_altas.append(
            Persona.objects.filter(created_at__year=year).count()
        )

    return JsonResponse({"reg": count_altas, "years": list(years)})


# def get_beneficiarios_activos(request, number):
#     data = None

#     with connection.cursor() as cursor:
#         query = (
#             f"SELECT COUNT(*) as ben_activo from (SELECT COUNT(adra_alimentos.persona_id) as ben_activos"  # noqa
#             f" FROM `adra_alimentos`  GROUP BY adra_alimentos.persona_id HAVING "  # noqa
#             f"COUNT(adra_alimentos.persona_id) >= {number}) as td"
#         )

#         cursor.execute(query)
#         data = cursor.fetchone()[0]
#     return JsonResponse({"num": data})


class CustomAllauthAdapter(DefaultAccountAdapter):
    """
    Clase encargada de enviar emails custom
    """

    def send_mail(self, template_prefix, email, context):
        tenant_info = get_tenant_model().objects.get(
            schema_name=connection.get_schema()
        )

        if context.get("activate_url"):
            user = context.get("user")

            sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
            message = Mail(
                from_email="admin@repartoalimentos.com",
                subject="Activación de la cuenta",
                to_emails=f"{user.email}",
            )
            message.dynamic_template_data = {
                "activate_url": f"{context.get('activate_url')}",
                "user": f"{context.get('user')}",
                "Sender_Name": f"{tenant_info.nombre}",
                "oar": f"{tenant_info.oar}",
                "Sender_Address": f"{tenant_info.calle}",
                "Sender_City": f"{tenant_info.ciudad}",
                "Sender_State": f"{tenant_info.provincia}",
                "Sender_Zip": f"{tenant_info.codigo_postal}",
            }
            message.template_id = str(
                settings.SENDGRID_ACTIVACION_CUENTA_TEMPLATE_ID
            )
            sg.send(message)

        elif context.get("password_reset_url"):
            user = context.get("user")
            sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
            message = Mail(
                from_email="admin@repartoalimentos.com",
                subject="Cambio de contraseña",
                to_emails=f"{user.email}",
            )
            message.dynamic_template_data = {
                "url_cambiar": f"{context.get('password_reset_url')}",
                "user": f"{user}",
                "Sender_Name": f"{tenant_info.nombre}",
                "oar": f"{tenant_info.oar}",
                "Sender_Address": f"{tenant_info.calle}",
                "Sender_City": f"{tenant_info.ciudad}",
                "Sender_State": f"{tenant_info.provincia}",
                "Sender_Zip": f"{tenant_info.codigo_postal}",
            }

            message.template_id = str(
                settings.SENDGRID_CHANGE_PASSWORD_TEMPLATE_ID
            )
            sg.send(message)


@login_required
def configuracion(request):
    excel_file = request.FILES.getlist("record", None)
    # print(excel_file)
    if len(excel_file):
        user = User.objects.get(id=request.user.pk)
        upload_payess = UploadExcelUsers(excel_file[0], user).upload_payees()
        return JsonResponse({"usuarios_fraud": list(upload_payess)})

    if request.method == "POST":
        print(request.POST)
        if "form_reparto" in request.POST.keys():
            pst = request.POST.copy()
            del pst["form_reparto"]
            alm_repatir = AlimentosRepatir.objects.all().first()

            form = AlimentosRepatrirForm(pst, instance=alm_repatir)
            # check whether it's valid:
            if form.is_valid():
                print("is valid")
                alm_rpt = form.save(commit=False)
                alm_rpt.modificado_por = request.user
                alm_rpt.save()
                messages.success(
                    request, "La configuraciòn se ha guardado correctamente "
                )
                return redirect("adra:configuracion")
            else:
                messages.error(
                    request,
                    "Formulario Invalido,revisar los datos introducidos ",
                )
                return redirect("adra:configuracion")

        if "reset_papeles" in request.POST:
            persona = Persona.objects.all()
            print("cambiar el estado de los papeles")
            for p in persona:
                p.empadronamiento = False
                p.libro_familia = False
                p.fotocopia_dni = False
                p.prestaciones = False
                p.nomnia = False
                p.cert_negativo = False
                p.aquiler_hipoteca = False
                p.recibos = False
                p.are_acte = False
                p.save()
            messages.success(request, "La tarea se ha relizado correctamente ")
            return redirect("adra:configuracion")

        else:
            form = DelegacionForm(request.POST)
            # check whether it's valid:
            if form.is_valid():
                print("es valido")

                nombre = form.cleaned_data["nombre"]
                oar = form.cleaned_data["oar"]
                codigo_postal = form.cleaned_data["codigo_postal"]
                ciudad = form.cleaned_data["ciudad"]
                calle = form.cleaned_data["calle"]
                provincia = form.cleaned_data["provincia"]
                # print(nombre, oar, codigo_postal, ciudad, calle, provincia)
                Delegaciones.objects.filter(pk=request.tenant.pk).update(
                    nombre=nombre,
                    oar=oar,
                    codigo_postal=codigo_postal,
                    ciudad=ciudad,
                    calle=calle,
                    provincia=provincia,
                )
                messages.success(
                    request, "La informacion se actualizado correctamente"
                )
                return redirect("adra:configuracion")

    else:
        delegaciones = Delegaciones.objects.get(pk=request.tenant.pk)
        delegacion_form = DelegacionForm(
            initial={
                "nombre": delegaciones.nombre,
                "oar": delegaciones.oar,
                "codigo_postal": delegaciones.codigo_postal,
                "ciudad": delegaciones.ciudad,
                "calle": delegaciones.calle,
                "provincia": delegaciones.provincia,
            }
        )

        alm_repatir = AlimentosRepatir.objects.all().first()
        init_reparto_alimento = {}
        for index in range(1, 17):
            init_reparto_alimento[f"alimento_{index}"] = getattr(
                alm_repatir, f"alimento_{index}"
            )
            init_reparto_alimento[f"alimento_{index}_type"] = getattr(
                alm_repatir, f"alimento_{index}_type"
            )
            init_reparto_alimento[f"alimento_{index}_0_3"] = getattr(
                alm_repatir, f"alimento_{index}_0_3"
            )
            init_reparto_alimento[f"alimento_{index}_4_6"] = getattr(
                alm_repatir, f"alimento_{index}_4_6"
            )
            init_reparto_alimento[f"alimento_{index}_7_9"] = getattr(
                alm_repatir, f"alimento_{index}_7_9"
            )

        alimentos_a_repatir_form = AlimentosRepatrirForm(
            initial=init_reparto_alimento
        )
        return render(
            request,
            "acciones/index.html",
            {
                "nbar": "acciones",
                "delegacion_form": delegacion_form,
                "alimentos_repatir": alimentos_a_repatir_form,
            },
        )


@login_required
def dashboard(request):
    beneficiar = Persona.objects.filter(active=True).exclude(covid=True)
    familiares = Hijo.objects.filter(
        persona__in=Persona.objects.filter(active=True).exclude(covid=True)
    )
    currentYear = datetime.now().year
    beneficiars_year = Persona.objects.filter(
        active=True, created_at__year=currentYear
    )
    last_five_benefiarios = Persona.objects.filter(active=True).order_by(
        "-created_at"
    )[:5]
    current_entregas = Alimentos.objects.all().count()
    ages = AgeCalculacion(beneficiar, familiares).calculate_age()

    return render(
        request,
        "adra/dashboard.html",
        {
            "nbar": "dashboard",
            "total_personas": ages["total_personas"],
            "beneficiars_year": len(beneficiars_year),
            "currentYear": currentYear,
            "last_five_benefiarios": last_five_benefiarios,
            "current_entregas": current_entregas,
            # "delegacion_form": delegacion_form,
            # "alimentos_repatir": alimentos_a_repatir_form,
        },
    )


@login_required
def generar_recuento_beneficiaros(request):
    res_pdf = io.BytesIO()
    val = RecuntoBeneficiarios().getRecuentoFile()
    val.write(res_pdf)
    res_pdf.seek(0)

    response = HttpResponse(
        res_pdf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # noqa
    )
    response[
        "Content-Disposition"
    ] = f"attachment; filename=recuento_beneficiarios.docx"
    # response['Content-Length'] = res_pdf.tell()

    return response
