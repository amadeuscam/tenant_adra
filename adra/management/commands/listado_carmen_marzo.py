from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Alignment
from adra.models import Persona


class Command(BaseCommand):

    def handle(self, *args, **options):

        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        # Delete the default worksheet
        # workbook.remove(workbook.active)
        worksheet.title = 'Beneficiarios'
        # Define the titles for columns
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 40
        worksheet.column_dimensions['D'].width = 40
        columns = [
            'Nombre',
            'Numero personas',
            'Domicilio',
            'Telefono',

        ]
        row_num = 1
        # fill = PatternFill(start_color='43fb32',
        #                    fill_type='solid')
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        persona = Persona.objects.all()
        for ben in persona:
            row_num += 1
            # Define the data for each cell in the row
            row = [
                ben.nombre_apellido,
                ben.hijo.all().count() + 1,
                ben.domicilio,
                ben.telefono,
            ]
            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                # cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

            # for d in ben.hijo.all().count():
            #     row_hijos = [
            #         '-',
            #         d.nombre_apellido,
            #         d.dni,
            #         '',
            #         d.fecha_nacimiento.strftime("%d/%m/%Y"),
            #     ]
            #     row_num += 1

            #     # Assign the data for each cell of the row
            #     for col_num, cell_value in enumerate(row_hijos, 1):
            #         cell = worksheet.cell(row=row_num, column=col_num)
            #         cell.value = cell_value
            #         cell.alignment = Alignment(horizontal='center')

        workbook.save("listado5.xlsx")
